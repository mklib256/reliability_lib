#Functions used in S6_Indices_calc

import numpy as np
import os
import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt
from scipy import stats

# Function to find the failures in a load column and return the failure indices and their duration
# Identifies failures based on load values in a specific load column (LP_LC).
# Returns the start indices of failures (f1) in the load column and their durations (dur).
def find_failures(LP_LC, threshold):
    failure_mask = LP_LC <= threshold  # Create a boolean array where failure is True
    f1 = []  # To store start indices of failures
    dur = [] # To store durations of failures
    failure_started = False
    start_index = None
    for i in range(len(failure_mask)): #For every failure, save the index when it starts and when it stops
        if failure_mask[i] and not failure_started:
            # Failure starts
            start_index = i
            failure_started = True
        elif not failure_mask[i] and failure_started:
            # Failure ends
            f1.append(start_index)
            dur.append(i - start_index)  # Duration is the difference between end and start indices for that failure
            failure_started = False
    
    # Handle case where failure lasts until the end of the time series
    if failure_started:
        f1.append(start_index) #failure start indices saved in f1
        dur.append(len(failure_mask) - start_index) #failure duration saved in dur
    return f1, dur

# Function to aggregate failures to each LP arising from MV and LV part
'''
def aggregate_LF_MV_LV(f1, dur, f1_LV, dur_LV, LP_LC):
    f1_LV = np.array(f1_LV, dtype=int)
    dur_LV = np.array(dur_LV)
    
    # Build the state matrix with the failed states from the MV part of the network
    LF_state = np.zeros(len(LP_LC))
    for start, duration in zip(f1, dur):
        end = start + duration
        LF_state[start:end] = 1

    # Check if any value in f1_LV conincides with a failed state in the MV part of the network and if so withdraw that value from f1_LV and dur
    LF_state = np.array(LF_state)
    mask = LF_state[f1_LV] >= 1 
    f1_double = f1_LV[mask]
    mask = ~np.isin(f1_LV, f1_double)
    f1_LV = f1_LV[mask]

    # Build the state matrix with the failed states from the LV part of the network
    LF_LV_state = np.zeros(len(LP_LC))
    for start, duration in zip(f1_LV, dur_LV):
        end = start + duration
        LF_LV_state[start:end] = 1

    # Unless there is a failed state in both matrices, add the failed states from both the MV and LV part into an aggregate
    # If there is a failed state in both matrices, consider the aggregate to have only one failure
    mask = ~((LF_state == 1) & (LF_LV_state == 1))
    LF_agg = np.where(mask, LF_state + LF_LV_state, 1)
    LF_agg = 1 - LF_agg  #change zeros to ones and ones to zeros
    indices = np.where(LF_agg == 1)[0]
    LF_agg[indices] = LP_LC[indices]
    
    return np.array(LF_agg) # Return the aggregate state matrix for the load point
'''

def aggregate_LF_MV_LV(f1, dur, f1_LV, dur_LV, LP_LC):
    # Convert inputs to NumPy arrays
    f1 = np.array(f1, dtype=int)
    dur = np.array(dur)
    f1_LV = np.array(f1_LV, dtype=int)
    dur_LV = np.array(dur_LV)
    LP_LC = np.array(LP_LC)
    n = len(LP_LC)

    # Build the state matrix for MV part
    LF_state = np.ones(n, dtype=int) #Load state matrix - Zero is for the OFF(Interrupted state) and 1 is for the ON state
    mask = np.zeros(n, dtype=bool) # Create a mask of indices for interrupted timesteps
    for start, duration in zip(f1, dur):
        mask[start:start + duration] = True
    LF_state[mask] = 0 # Set all interrupted timesteps to 0

    #Deal with double interruptions - interruptions in the LV part that occur when the MV part is already interrupted are eliminated
    # Remove interruptions in the LV network that start during MV network interruptions
    mask = LF_state[f1_LV] == 1 #Identify only the interruptions in the LV part where there is no interruption in the MV part
    f1_LV = f1_LV[mask] #Update f1_LV with only these interruptions
    dur_LV = dur_LV[mask] #Update the durations for only these interruptions

    # Build the state matrix for LV part
    LF_LV_state = np.ones(n, dtype=int)
    mask = np.zeros(n, dtype=bool)
    for start, duration in zip(f1_LV, dur_LV):
        mask[start:start + duration] = True
    LF_LV_state[mask] = 0  # Set all interrupted timesteps to 0

    #Deal with overlapping interruptions - interruptions that occur from both the MV and LV part at the same time are merged
    LF_agg_state = np.logical_and(LF_state, LF_LV_state).astype(int) # Combine MV and LV states
    LF_agg = np.where(LF_agg_state == 1, LP_LC, 0)  # Apply LP_LC values

    return LF_agg
'''
def update_f1_LV_dur_LV(f1, dur, f1_LV, dur_LV, LP_LC):
    # Convert inputs to NumPy arrays
    f1 = np.array(f1, dtype=int)
    dur = np.array(dur)
    f1_LV = np.array(f1_LV, dtype=int)
    dur_LV = np.array(dur_LV)
    LP_LC = np.array(LP_LC)
    n = len(LP_LC)

    # Build the state matrix for MV part
    LF_state = np.ones(n, dtype=int) #Load state matrix - Zero is for the OFF(Interrupted state) and 1 is for the ON state
    mask = np.zeros(n, dtype=bool) # Create a mask of indices for interrupted timesteps
    for start, duration in zip(f1, dur):
        mask[start:start + duration] = True
    LF_state[mask] = 0 # Set all interrupted timesteps to 0

    # Build the state matrix for LV part
    LF_LV_state = np.ones(n, dtype=int) #Load state matrix - Zero is for the OFF(Interrupted state) and 1 is for the ON state
    mask = np.zeros(n, dtype=bool) # Create a mask of indices for interrupted timesteps
    for start, duration in zip(f1_LV, dur_LV):
        mask[start:start + duration] = True
    LF_LV_state[mask] = 0 # Set all interrupted timesteps to 0

    #Deal with double interruptions - interruptions in the LV part that occur when the MV part is already interrupted are eliminated
    # Remove interruptions in the LV network that start during MV network interruptions
    mask = LF_state[f1_LV] == 1 #Identify only the interruptions in the LV part where there is no interruption in the MV part
    f1_LV = f1_LV[mask] #Update f1_LV with only these interruptions
    dur_LV = dur_LV[mask] #Update the durations for only these interruptions

    return f1_LV, dur_LV
'''
def update_f1_LV_dur_LV(f1, dur, f1_LV, dur_LV, LP_LC):
    # Convert inputs to NumPy arrays
    f1 = np.array(f1, dtype=int)
    dur = np.array(dur)
    f1_LV = np.array(f1_LV, dtype=int)
    dur_LV = np.array(dur_LV)
    LP_LC = np.array(LP_LC)
    n = len(LP_LC)

    # Build the state matrix for MV part (LF_state)
    LF_state = np.ones(n, dtype=int)
    mask = np.zeros(n, dtype=bool)
    for start, duration in zip(f1, dur):
        mask[start:start + duration] = True
    LF_state[mask] = 0

    # Build the state matrix for LV part (LF_LV_state)
    LF_LV_state = np.ones(n, dtype=int)
    mask = np.zeros(n, dtype=bool)
    for start, duration in zip(f1_LV, dur_LV):
        mask[start:start + duration] = True
    LF_LV_state[mask] = 0

    for i in range (n):
        if LF_state[i] == 0:
            LF_LV_state[i] = 0
            
    '''
    # Update f1_LV and dur_LV based on the modified LF_LV_state
    new_f1_LV = []
    new_dur_LV = []
    in_interruption = False
    start = 0
    for i in range(n):
        if LF_LV_state[i] == 0 and not in_interruption:
            in_interruption = True
            start = i
        elif LF_LV_state[i] == 1 and in_interruption:
            in_interruption = False
            new_f1_LV.append(start)
            new_dur_LV.append(i - start)

    if in_interruption:  # Handle case where interruption lasts until the end
        new_f1_LV.append(start)
        new_dur_LV.append(n - start)
    '''
    return LF_LV_state

# Function to calculate ENS in a year based on the failure indices f1 and their duration
def calculate_ens(f1, dur, LP_LC, dt):
    ENS = []  # To store ENS values for each failure

    for m in range(len(f1)):
        if dur[m] >= 1:
            # Step 1: Extract power values for the failure duration
            D = np.zeros(dur[m])  # Initialize power usage array
            r = 0  # Initialize row index for D
            for p in range(f1[m], f1[m] + dur[m]):  # Loop through the failure period
                D[r] = LP_LC[p]  # Extract power from LP matrix at time p and load index h
                r += 1

            # Step 2: Apply the trapezoidal rule to calculate ENS
            E = np.zeros(len(D) - 1)  # Initialize ENS array
            for q in range(len(D) - 1):
                E[q] = 0.5 * dt * (D[q] + D[q + 1])  # Trapezoidal formula

            # Step 3: Sum the ENS for this failure event
            ENS.append(np.sum(E))

    return np.array(ENS)

# Function to create the PDF and CDF plots (smoothened)
def pdf_cdf_plot_smooth(Index_annual, index_name, target_folder):
    data = Index_annual.flatten()  # Flatten the matrix into a 1D array

    # Create PDF using KDE estimator for smoothening
    kde = stats.gaussian_kde(data)
    x_range = np.linspace(min(data), max(data), 10000)
    smoothened_pdf = kde(x_range)
    pdf = smoothened_pdf /len(data)

    # Create CDF
    cdf = np.cumsum(pdf) / np.sum(pdf)
    #cdf = np.cumsum(pdf) / np.sum(pdf)

    # Plotting
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))

    # PDF plot
    ax1.plot(x_range, pdf)
    ax1.set_title(f'Probability Density Function (PDF) of {index_name}')
    ax1.set_xlabel(index_name)
    ax1.set_ylabel('Density')

    # CDF plot
    ax2.plot(x_range, cdf)
    ax2.set_title(f'Cumulative Distribution Function (CDF) of {index_name}')
    ax2.set_xlabel(index_name)
    ax2.set_ylabel('Cumulative Probability')

    plt.tight_layout()

    # Save the figure
    filename = f'{index_name}_PDF_CDF_smooth.png'
    plt.savefig(os.path.join(target_folder, filename), dpi=300, bbox_inches='tight')
    plt.close()  # Close the figure to free up memory

# Function to create the PDF and CDF plots (detailed)
def pdf_cdf_plot_detailed(Index_annual, index_name, target_folder):
    data = Index_annual.flatten()  # Flatten the matrix into a 1D array
    metric = np.round(data,2) #Round to one decimal place

    # Determine the range for plotting
    step = 0.01
    max_x = np.max(metric)
    plotvalues = np.arange(0, max_x + step, step)

    # Calculate PDF
    vals, bin_edges = np.histogram(metric, bins=plotvalues, density=True)
    Index_PDF = vals / len(metric)
    Index_PDF = Index_PDF.reshape(-1, 1)

    # Calculate CDF
    Index_CDF = np.cumsum(Index_PDF) / np.sum(Index_PDF)

    # Plotting
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))

    # PDF plot
    ax1.plot(plotvalues[:-1], Index_PDF, label='PDF')
    ax1.set_title(f'Probability Density Function (PDF) of {index_name}')
    ax1.set_xlabel(index_name)
    ax1.set_ylabel('Density')
    ax1.legend()

    # CDF plot
    ax2.plot(plotvalues[:-1], Index_CDF, label='CDF')
    ax2.set_title(f'Cumulative Distribution Function (CDF) of {index_name}')
    ax2.set_xlabel(index_name)
    ax2.set_ylabel('Cumulative Probability')

    plt.tight_layout()

    # Save the figure
    filename = f'{index_name}_PDF_CDF_detailed.png'
    plt.savefig(os.path.join(target_folder, filename), dpi=300, bbox_inches='tight')
    plt.close()  # Close the figure to free up memory


# Function to save the results in the text file to an excel sheet
def copy_text_to_excel(text_file_path, excel_file_path):
    # Check if the Excel file already exists
    if os.path.exists(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)  # If it exists, load the workbook

        if 'Python' in workbook.sheetnames:  # Check if 'Python' sheet already exists
            sheet = workbook['Python']
            sheet.delete_rows(1, sheet.max_row)  # Clear existing content
        else:
            sheet = workbook.create_sheet('Python')  # Create a new 'Python' sheet
    else:
        workbook = Workbook()  # If it doesn't exist, create a new workbook
        workbook.remove(workbook.active)  # Remove the default sheet
        sheet = workbook.create_sheet('Python')  # Create the 'Python' sheet

    # Open and read the text file
    with open(text_file_path, 'r') as text_file:
        for row, line in enumerate(text_file, start=1):
            columns = line.strip().split('\t')  # Split the line into columns (assuming tab-separated values)
            for col, value in enumerate(columns, start=1):
                sheet.cell(row=row, column=col, value=value)  # Write each column to the Excel sheet

    # Save the Excel file
    workbook.save(excel_file_path)
    print(f"Results have been saved in Results.xlsx")



